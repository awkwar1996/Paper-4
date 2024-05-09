'''
解析MIP结果，得到数据
'''
import openpyxl
import os
from itertools import product
def analysis_result(model, N, B):
    solved_instances_num = 0
    solved_instances_time = 0
    solved_instances_value = 0
    instance_time = 0
    instances_value = 0
    instances_num = 0

    url = f'./Result/model_result/{model}/B{B}'
    fires = os.listdir(url)
    for fire in fires:
        wb = openpyxl.load_workbook(url+'/'+fire)
        basic_data = wb[wb.sheetnames[0]]
        if int(basic_data.cell(row=2, column=1).value)*int(basic_data.cell(row=2, column=2).value) == N:# S*H=N
            instances_num += 1
            instance_time += min(3600, basic_data.cell(row=2, column=7).value)
            instances_value += basic_data.cell(row=2, column=5).value
            if int(basic_data.cell(row=2, column=4).value) == 2:# status=2
                solved_instances_num += 1
                solved_instances_time += min(3600, basic_data.cell(row=2, column=7).value)
                solved_instances_value += basic_data.cell(row=2, column=5).value
    instance_time = instance_time/instances_num
    instances_value = instances_value/instances_num
    if solved_instances_num != 0:
        solved_instances_time = solved_instances_time/solved_instances_num
        solved_instances_value = solved_instances_value/solved_instances_num
    return instances_num, round(solved_instances_num/instances_num, 2), round(solved_instances_value, 2), round(instances_value,2), round(solved_instances_time,2), round(instance_time,2)
def analysis_total_result(model):
    solved_instances_num = 0
    solved_instances_time = 0
    solved_instances_value = 0
    instance_time = 0
    instances_value = 0
    instances_num = 0

    url = f'./Result/model_result/{model}'
    batch_fires = os.listdir(url)
    for batch_fire in batch_fires:
        fires = os.listdir(url+'/'+batch_fire)
        for fire in fires:
            wb = openpyxl.load_workbook(url + '/' + batch_fire+'/'+fire)
            basic_data = wb[wb.sheetnames[0]]
            instances_num += 1
            instance_time += min(3600, basic_data.cell(row=2, column=7).value)
            instances_value += basic_data.cell(row=2, column=5).value
            if int(basic_data.cell(row=2, column=4).value) == 2:  # status=2
                solved_instances_num += 1
                solved_instances_time += min(3600, basic_data.cell(row=2, column=7).value)
                solved_instances_value += basic_data.cell(row=2, column=5).value
    instance_time = instance_time / instances_num
    instances_value = instances_value / instances_num
    if solved_instances_num != 0:
        solved_instances_time = solved_instances_time / solved_instances_num
        solved_instances_value = solved_instances_value / solved_instances_num
    return instances_num, round(solved_instances_num / instances_num, 2), round(solved_instances_value, 2), round(
        instances_value, 2), round(solved_instances_time, 2), round(instance_time, 2)
if __name__ == '__main__':
    # model_list = ['BIMV']
    # N_list = [60]
    # B_list = [1]
    # for model, N, B in product(model_list, N_list, B_list):
    #     print(f'model={model}, N={N}, B={B},',end='')
    #     instances_num, solved_instances_percent, solved_instances_value, instances_value, solved_instances_time, instances_time = analysis_result(model, N, B)
    #
    #     print('instance_num = %d\n#opt.=%.2f, f^{opt.}=%.2f, f = %.2f, T^{opt.}=%.2f, T=%.2f\n'%(instances_num, solved_instances_percent, solved_instances_value, instances_value, solved_instances_time, instances_time))

    model_list = ['BIMV']
    for model in model_list:
        print(f'model={model},', end='')
        instances_num, solved_instances_percent, solved_instances_value, instances_value, solved_instances_time, instances_time = analysis_total_result(
            model)
        print('instance_num = %d\n#opt.=%.2f, f^{opt.}=%.2f, f = %.2f, T^{opt.}=%.2f, T=%.2f\n' % (
        instances_num, solved_instances_percent, solved_instances_value, instances_value, solved_instances_time,
        instances_time))

