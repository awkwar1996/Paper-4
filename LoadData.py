#!/usr/bin/env python
from itertools import product

import numpy as np
import openpyxl
import pandas as pd
'''
Loading the data
Notice: plate in the lower are outbounded first. //excel 下方的先出库
'''

def LoadData(url):
    print('----------loading data----------')
    basicSheet = openpyxl.load_workbook(url)['Basic']
    S = int(basicSheet.cell(2, 1).value)
    H = int(basicSheet.cell(2, 2).value)

    plateData = pd.read_excel(url, sheet_name=1)
    #print(plateData)
    N = plateData.shape[0]
    P_max = int(plateData.max()['group'])
    L_max = plateData.max()['long']
    L_min = plateData.min()['long']

    inbound_config = [[] for i in range(int(plateData.max()['batch']))]

    for index, plate_info in plateData.iterrows():
        inbound_config[int(plate_info['batch']) - 1].append(index)

    # 初始
    O = dict()
    L = dict()
    P = dict()

    P_s = np.zeros((N, N))
    L_s = np.zeros((N, N))

    for index in range(N):
        temp_O = []
        temp_L = []
        temp_P = []
        for anotherIndex in range(N):
            L_s[index][anotherIndex] = max(0, plateData.loc[index, 'long'] - plateData.loc[anotherIndex, 'long'])
            P_s[index][anotherIndex] = max(0, plateData.loc[index, 'group'] - plateData.loc[anotherIndex, 'group'])
            if index == anotherIndex: continue
            # 1. 判断长度，小的加入L
            if plateData.loc[index, 'long'] > plateData.loc[anotherIndex, 'long']: temp_L.append(anotherIndex)
            # 2. 判断分组，小的加入P
            if plateData.loc[index, 'group'] > plateData.loc[anotherIndex, 'group']: temp_P.append(anotherIndex)
            # 3. 判断是否在同一个批次，同一个批次的再判断顺序
            if plateData.loc[index, 'batch'] == plateData.loc[anotherIndex, 'batch'] and inbound_config[
                plateData.loc[index, 'batch'] - 1].index(index) < inbound_config[
                plateData.loc[index, 'batch'] - 1].index(anotherIndex):
                temp_O.append(anotherIndex)
            L[index] = temp_L
            P[index] = temp_P
            O[index] = temp_O
    print('----------data loaded----------')
    return N, S, H, P_max, L_max, L_min, O, P, L, P_s, L_s, plateData, inbound_config
def O_plus(O):
    O_plus = {}
    for key in O.keys():
        O_plus[key] = []
        for another_key in O.keys():
            if another_key == key: continue
            if key in O[another_key]: O_plus[key].append(another_key)
    return O_plus
if __name__ == '__main__':
    N, S, H, P, _, _, O, _, _, _, _, plate_info, inbound_config = LoadData('DataFormat30.xlsx')
    O_plus = O_plus(O)
    print('hello')

