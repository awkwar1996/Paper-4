#!/usr/bin/env python
'''
Saving model solutions to excel
'''
import numpy as np
import openpyxl
import pandas as pd


def save_model_result_to_excel(model_status, result, optimized_storage, plate_data, S, H, P, Url, run_time):
    wb = openpyxl.Workbook()
    # sheet1 result information
    ws1 = wb.active
    ws1.cell(row=1, column=1, value='S')
    ws1.cell(row=2, column=1, value=S)
    ws1.cell(row=1, column=2, value='H')
    ws1.cell(row=2, column=2, value=H)
    ws1.cell(row=1, column=3, value='P')
    ws1.cell(row=2, column=3, value=P)
    ws1.cell(row=1, column=4, value='model status')
    ws1.cell(row=2, column=4, value=model_status)
    ws1.cell(row=1, column=5, value='relocation')
    ws1.cell(row=2, column=5, value=result)
    ws1.cell(row=1, column=6, value='long')
    ws1.cell(row=2, column=6, value='Null')
    ws1.cell(row=1, column=7, value='run time')
    ws1.cell(row=2, column=7, value=run_time)
    #sheet2 storage configuration
    ws2 = wb.create_sheet()
    if sum([sum(optimized_storage[i]) for i in range(S)]) > 0:
        for i, stack in enumerate(optimized_storage):
            ws2.cell(row=i + 1, column=1, value=len(stack))
            for j, plate in enumerate(stack):
                ws2.cell(row=i + 1, column=j + 2, value=plate)
    else: ws2.cell(row=1, column=1, value='Null')
    #sheet3 plate data
    ws3 = wb.create_sheet()
    ws3.cell(row=1, column=1, value='id')
    ws3.cell(row=1, column=2, value='group')
    ws3.cell(row=1, column=3, value='long')
    ws3.cell(row=1, column=4, value='batch')
    for index, row in plate_data.iterrows():
        ws3.cell(row=index + 2, column=1, value =index)
        ws3.cell(row=index + 2, column=2, value=row['group'])
        ws3.cell(row=index + 2, column=3, value=row['long'])
        ws3.cell(row=index + 2, column=4, value=row['batch'])
    wb.save(Url)

def save_algorithm_result_to_excel(instance, relocation, function, Url):
    wb = openpyxl.load_workbook(Url)
    ws = wb.active
    row = ws.max_row
    ws.cell(row=row + 1, column=1, value=instance)
    ws.cell(row=row + 1, column=2, value=relocation)
    ws.cell(row=row + 1, column=3, value=function)
    wb.save(Url)
def save_rule_algorithm_result_to_excel(instance, alg_result, Url):
    wb = openpyxl.load_workbook(Url)
    ws = wb.active
    row = ws.max_row
    ws.cell(row=row + 1, column=1, value=instance)
    ws.cell(row=row + 1, column=2, value=alg_result['FewestBlockageHeuristic'][0])
    ws.cell(row=row + 1, column=3, value=alg_result['FewestBlockageHeuristic'][1])
    ws.cell(row=row + 1, column=4, value=alg_result['LeastFilledStackHeuristic'][0])
    ws.cell(row=row + 1, column=5, value=alg_result['LeastFilledStackHeuristic'][1])
    ws.cell(row=row + 1, column=6, value=alg_result['MostSimilarHeuristic'][0])
    ws.cell(row=row + 1, column=7, value=alg_result['MostSimilarHeuristic'][1])
    ws.cell(row=row + 1, column=8, value=alg_result['FirstFitHeuristic'][0])
    ws.cell(row=row + 1, column=9, value=alg_result['FirstFitHeuristic'][1])
    ws.cell(row=row + 1, column=10, value=alg_result['BestFitHeuristic'][0])
    ws.cell(row=row + 1, column=11, value=alg_result['BestFitHeuristic'][1])

    wb.save(Url)
def save_SA_algorithm_result_to_excel(instance, relocation, Url):
    wb = openpyxl.load_workbook(Url)
    ws = wb.active
    row = ws.max_row
    ws.cell(row=row + 1, column=1, value=instance)
    ws.cell(row=row + 1, column=2, value=relocation)

    wb.save(Url)
if __name__ == '__main__':
    pass